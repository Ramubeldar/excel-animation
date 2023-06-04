# importing important python libraries
import moviepy.editor as mp
from openpyxl import load_workbook
import os
import cv2
import numpy as np


# defining the function
def read_excel_row(row_num, excel_file):
    wb = load_workbook(filename=excel_file)
    worksheet = wb.active

    Input_folder_name = []
    for col in range(1, 2):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Input_folder_name.append(cell_value)

    Message = []
    for col in range(2, 3):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Message.append(cell_value)

    Output_folder_name = []
    for col in range(3, 4):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Output_folder_name.append(cell_value)

    # Split the single string element into a list of separate elements
    Message = Message[0].split('\n')

    # Create a new list to hold the elements without newline character
    Messages = []

    # Loop through the original list, strip leading and trailing whitespaces and append the elements to the new list
    for element in Message:
        Messages.append(element.strip())

    # Define the path to the folder containing the input images
    # image_folder = "input_folder"
    image_folder = str(Input_folder_name[0])

    # Get the list of image files in the folder
    image_files = os.listdir(image_folder)

    # Read the input images
    images = []
    for file in image_files:
        image_path = os.path.join(image_folder, file)
        image = cv2.imread(image_path)
        if image is None:
            print("Failed to load image:", image_path)
            continue
        images.append(image)

    # Define the total duration of the video in seconds
    video_duration = 20

    # Set the frames per second (FPS) for the video
    fps = 30

    # Calculate the total number of frames for the video
    num_frames = video_duration * fps

    # Determine the output video dimensions
    output_width = 1280
    output_height = 720

    # Create a VideoWriter object
    output_folder = Output_folder_name[0]
    os.makedirs(output_folder, exist_ok=True)
    output_path = os.path.join(output_folder, "output.mp4")
    output = cv2.VideoWriter(output_path, cv2.VideoWriter_fourcc(*'mp4v'), fps, (output_width, output_height))

    # Define the motion animations
    animations = [
        (0, 0, output_width - images[0].shape[1], output_height - images[0].shape[0]),# Animation 1: Move to bottom right
        (output_width - images[1].shape[1], 0, 0, output_height - images[1].shape[0]),  # Animation 2: Move to top left
        (0, output_height - images[2].shape[0], output_width - images[2].shape[1], 0),  # Animation 3: Move to top right
        (output_width - images[3].shape[1], output_height - images[3].shape[0], 0, 0)# Animation 4: Move to bottom left
    ]

    # Calculate the number of frames for each image animation
    frames_per_animation = int(num_frames / (len(images)))  # Increase the divisor to slow down the animation

    # Generate the frames with motion animations for each image
    for image_idx, image in enumerate(images):
        # Select the animation for the current image
        animation_params = animations[image_idx % len(animations)]

        # Calculate the motion parameters
        start_x, start_y, end_x, end_y = animation_params
        dx = (end_x - start_x) / frames_per_animation
        dy = (end_y - start_y) / frames_per_animation

        # Generate the frames for the current image animation
        for frame_idx in range(frames_per_animation):
            # Calculate the current position of the image
            current_x = int(start_x + frame_idx * dx)
            current_y = int(start_y + frame_idx * dy)

            # Create an empty canvas with the output resolution
            output_frame = np.zeros((output_height, output_width, 3), dtype=np.uint8)

            # Calculate the top-left corner coordinates for pasting the image
            paste_x = max(0, current_x)
            paste_y = max(0, current_y)

            # Calculate the bottom-right corner coordinates for pasting the image
            paste_width = min(output_width - paste_x, image.shape[1])
            paste_height = min(output_height - paste_y, image.shape[0])

            # Calculate the top-left corner coordinates for cropping the image
            crop_x = max(0, -current_x)
            crop_y = max(0, -current_y)

            # Crop the image
            cropped_image = image[crop_y:crop_y + paste_height, crop_x:crop_x + paste_width]

            # Paste the cropped image at the current position onto the canvas
            output_frame[paste_y:paste_y + paste_height, paste_x:paste_x + paste_width] = cropped_image

            # Add the text to the center of the output frame
            text = Messages[image_idx]
            text_color = (255, 255, 255)  # White color
            text_thickness = 2
            text_font = cv2.FONT_HERSHEY_SIMPLEX
            text_size = cv2.getTextSize(text, text_font, 1, text_thickness)[0]
            text_x = int((output_width - text_size[0]) / 2)
            text_y = int((output_height + text_size[1]) / 2)
            cv2.putText(output_frame, text, (text_x, text_y), text_font, 1, text_color, text_thickness, cv2.LINE_AA)

            # Write the output frame to the video
            output.write(output_frame)

    # Release the VideoWriter object
    output.release()

    print("Output video saved:", output_path)


# reading the exel file
excel_file_sheets = "data.xlsx"
workbook = load_workbook("data.xlsx")

# Select the active worksheet
worksheet = workbook.active

# Find the number of rows with data in excel sheet
num_rows = 0
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
    if any(cell.value for cell in row):
        num_rows += 1

for row_number in range(2, num_rows + 1):
    read_excel_row(row_number, excel_file_sheets)
